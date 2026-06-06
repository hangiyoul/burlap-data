"""World Bank Pink Sheet — 월간 커피 가격(Arabica / Robusta).

신뢰성 1순위 공공 데이터. 1960년~ 매월 첫째 주 갱신.
WB 가 Pink Sheet 페이지에서 직접 링크하는 XLSX 의 'Monthly Prices' 시트:
  - 행5: 헤더 ("Coffee, Arabica", "Coffee, Robusta")
  - 행6: 단위 ($/kg)
  - 행7~: 데이터. 컬럼1 = 'YYYYMmm' (예: '2026M05')
  - 컬럼13 = Arabica, 컬럼14 = Robusta

단위 변환:
  - Arabica: $/kg → ¢/lb  (× 45.3592)
  - Robusta: $/kg → $/t   (× 1000)

이 어댑터는 history 를 REPLACE 합니다 (WB 가 전체 시계열을 제공하므로 누적 불필요).
URL 은 매년 바뀌므로 worldbank.org/en/research/commodity-markets 페이지에서 동적 추출.
"""
import re
import urllib.request
from io import BytesIO

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
LISTING_URL = "https://www.worldbank.org/en/research/commodity-markets"


def _fetch(url, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def _find_xlsx_url() -> str:
    """WB 페이지에서 'CMO-Historical-Data-Monthly.xlsx' 최신 링크를 추출.
    파일명에 연도 코드가 들어있어서 매년 바뀜.
    """
    html = _fetch(LISTING_URL, timeout=30).decode("utf-8", "ignore")
    m = re.search(r'(https?://[^"]*CMO-Historical-Data-Monthly\.xlsx)', html)
    if not m:
        raise RuntimeError("WB Pink Sheet XLSX URL not found on listing page")
    return m.group(1)


def _parse_yyyy_m(s: str):
    """'2026M05' → (2026, 5)"""
    m = re.match(r"(\d{4})M(\d{1,2})", str(s))
    return (int(m.group(1)), int(m.group(2))) if m else None


def fetch_coffee_monthly():
    """반환:
      {
        "arabica": {"dates": ["1960-01-01", ...], "values": [42.6, ...], "unit": "¢/lb"},
        "robusta": {"dates": [...], "values": [...], "unit": "$/t"},
        "source_url": "...",
      }
    """
    from openpyxl import load_workbook   # 지연 임포트

    xlsx_url = _find_xlsx_url()
    data = _fetch(xlsx_url, timeout=60)
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb["Monthly Prices"]

    # 헤더 행에서 Arabica·Robusta 컬럼 동적 탐색 (위치 변경 대비)
    header_row = 5
    ar_col = ro_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if not v:
            continue
        s = str(v).strip().lower()
        if s == "coffee, arabica":
            ar_col = c
        elif s == "coffee, robusta":
            ro_col = c
    if not ar_col or not ro_col:
        raise RuntimeError(f"Coffee columns not found (ar={ar_col}, ro={ro_col})")

    ar_dates, ar_vals = [], []
    ro_dates, ro_vals = [], []
    for r in range(7, ws.max_row + 1):
        date_raw = ws.cell(r, 1).value
        ym = _parse_yyyy_m(date_raw) if date_raw else None
        if not ym:
            continue
        date_iso = f"{ym[0]:04d}-{ym[1]:02d}-01"
        ar = ws.cell(r, ar_col).value
        ro = ws.cell(r, ro_col).value
        if isinstance(ar, (int, float)):
            ar_dates.append(date_iso)
            ar_vals.append(round(ar * 45.3592, 2))   # $/kg → ¢/lb
        if isinstance(ro, (int, float)):
            ro_dates.append(date_iso)
            ro_vals.append(round(ro * 1000.0, 0))    # $/kg → $/t

    if not ar_vals or not ro_vals:
        raise RuntimeError("No coffee data parsed")

    return {
        "arabica": {"dates": ar_dates, "values": ar_vals, "unit": "¢/lb"},
        "robusta": {"dates": ro_dates, "values": ro_vals, "unit": "$/t"},
        "source_url": xlsx_url,
    }


if __name__ == "__main__":
    r = fetch_coffee_monthly()
    print(f"Arabica: {len(r['arabica']['values'])} months, "
          f"{r['arabica']['dates'][0]} → {r['arabica']['dates'][-1]}, "
          f"last = {r['arabica']['values'][-1]} {r['arabica']['unit']}")
    print(f"Robusta: {len(r['robusta']['values'])} months, "
          f"{r['robusta']['dates'][0]} → {r['robusta']['dates'][-1]}, "
          f"last = {r['robusta']['values'][-1]} {r['robusta']['unit']}")
    print(f"Source: {r['source_url']}")
